import datetime
from io import BytesIO
import frappe
from frappe.model.document import Document
from frappe.utils import add_days
from frappe.utils.file_manager import get_file
import openpyxl
from openpyxl.utils import column_index_from_string


TAG_NAME_HEADER_TEXT = "Tag Name"

EXCEL_FILES_FIELDNAME = "dmr_boiler_excel_upload"

DEFAULT_MIN_MAX_AVG_COLUMNS = {
	"engg_units_col": "E",
	"max_col": "F",
	"max_time_col": "G",
	"min_col": "H",
	"min_time_col": "I",
	"avg_col": "J",
}

DEFAULT_MIN_MAX_AVG_ROWS = {
	"engg_units_row": 2,
	"max_row": 3,
	"max_time_row": 4,
	"min_row": 5,
	"min_time_row": 6,
	"avg_row": 7,
}

PLANT_CONFIG = {
	"RSL Louhka": {
		"sources": {
			"Boiler Report": {
				"tag_name_col": "B",
				"hourly_start_col": "K",
				"hourly_end_col": "AH",
				"next_day_start_col": None,
				"next_day_end_col": None,
				"field_tag_map": {
					"main_steam_pressure": {"tag": "PT302", "label": "Main Steam Pressure", "agg": "avg"},
					"main_steam_temprature": {"tag": "TT303", "label": "Main Steam Temprature", "agg": "avg"},
					"float_reke": {"tag": "TE414", "label": "ESP Outlet Temp", "agg": "avg"},
					"oxygen__at_eco_ol": {"tag": "AT401", "label": "Oxygen % At Eco O/L", "agg": "avg"},
					"boiler_feed_water_flow": {"tag": "FT301", "label": "Boiler Feed Water Flow", "agg": "sum"},
					"float_zcpn": {"tag": "FT302", "label": "Steam Produced", "agg": "sum"},
					"dm_flow_to_dearator": {"tag": "FT101", "label": "DM Flow To Dearator ", "agg": "sum"},
					"float_pvrh": {"tag": "EKW2000", "label": "Power Generation", "agg": "sum"},
					"turbine_steam": {"tag": "FT2001", "label": "Turbine Steam", "agg": "sum"},
					"deaerator": {"tag": "FT102", "label": "Deartor", "agg": "sum"},
					"turbine_chest_pressure":{"tag": "ACT1000", "label": "Turbine Chest Pressure", "agg": "sum"},
				},
			},
   
   
   	"Evaporator Sheet": {
				"orientation": "column",
				"tag_name_row": 3,
				"hourly_start_row": 4,
				"hourly_end_row": 27,
				"next_day_start_row": None,
				"next_day_end_row": None,
				"min_max_avg_rows": {
					"engg_units_row": 29,
					"max_row": 30,
					"max_time_row": 31,
					"min_row": 32,
					"min_time_row": 33,
					"avg_row": 34,
				},
				"field_tag_map": {
					"evaporation": {
						"label": "Evaporation",
						"formula": [
							{"col": "B", "row": 29},
							{"col": "C", "row": 29},
						],
					}
				},
			},
		},
	},
	"Superior Biofuels": {
		"sources": {
			"Boiler Report": {
				"tag_name_col": "B",
				"hourly_start_col": "K",
				"hourly_end_col": "AH",
				"next_day_start_col": None,
				"next_day_end_col": None,
				"field_tag_map": {
					"main_steam_pressure": {"tag": "PT302", "label": "Main Steam Pressure", "agg": "avg"},
					"main_steam_temprature": {"tag": "TT306", "label": "Main Steam Temprature", "agg": "avg"},
					"float_reke": {"tag": "TE415", "label": "ESP Outlet Temp", "agg": "avg"},
					"oxygen__at_eco_ol": {"tag": "AT401", "label": "Oxygen % At Eco O/L", "agg": "avg"},
					"boiler_feed_water_flow": {"tag": "FT301", "label": "Boiler Feed Water Flow", "agg": "sum"},
					"float_zcpn": {"tag": "FT302", "label": "Steam Produced", "agg": "sum"},
					"dm_flow_to_dearator": {"tag": "FI_2014", "label": "DM Flow To Dearator ", "agg": "sum"}
				},
			},
		},
	},

	"RSL Belwara": {
		"sources": {
			"Boiler Report": {
				"tag_name_col": "B",
				"hourly_start_col": "K",
				"hourly_end_col": "AH",
				"next_day_start_col": None,
				"next_day_end_col": None,
				"field_tag_map": {
					"main_steam_pressure": {"tag": "PT302", "label": "Main Steam Pressure", "agg": "avg"},
					"main_steam_temprature": {"tag": "TT304", "label": "Main Steam Temprature", "agg": "avg"},
					"float_reke": {"tag": "TE417", "label": "ESP Outlet Temp", "agg": "avg"},
					"oxygen__at_eco_ol": {"tag": "AT401", "label": "Oxygen % At Eco O/L", "agg": "avg"},
					"boiler_feed_water_flow": {"tag": "FT201", "label": "Boiler Feed Water Flow", "agg": "sum"},
					"float_zcpn": {"tag": "FT301", "label": "Steam Produced", "agg": "sum"},
					"dm_flow_to_dearator": {"tag": "FI_2014", "label": "DM Flow To Dearator ", "agg": "sum"},
					"deaerator": {"tag": "FI_2013", "label": "Deaerator", "agg": "sum"},
     				"feed_water_inlet_temp": {"tag": "TE201", "label": "Feed Water Inlet Temperature", "agg": "avg"},
              		"return_codensate": {"tag": "FT102", "label": "Return Condensate", "agg": "sum"},
              		"exhaust_pressure": {"tag": "PT01", "label": "Exhaust Pressure", "agg": "sum"}
				},
			},
			"Distillary 80 KLPD": {
				"orientation": "column",
				"tag_name_row": 3,
				"hourly_start_row": 4,
				"hourly_end_row": 27,
				"next_day_start_row": None,
				"next_day_end_row": None,
				"min_max_avg_rows": {
					"engg_units_row": 29,
					"max_row": 30,
					"max_time_row": 31,
					"min_row": 32,
					"min_time_row": 33,
					"avg_row": 34,
				},
				"field_tag_map": {
					"liquification": {
						"label": "Liquification",
						"formula": [
							{"col": "T", "row": 33},
							{"col": "U", "row": 33},
						],
					},
					"distillation": {
						"label": "Distillation",
						"formula": [
							{"col": "O", "row": 33, "scale": 1000},
							{"col": "P", "row": 33},
						],
					},
				},
			},
   
   			"Distillary 300 KLPD": {
					"orientation": "column",
					"tag_name_row": 3,
					"hourly_start_row": 4,
					"hourly_end_row": 27,
					"next_day_start_row": None,
					"next_day_end_row": None,
					"min_max_avg_rows": {
						"engg_units_row": 29,
						"max_row": 30,
						"max_time_row": 31,
						"min_row": 32,
						"min_time_row": 33,
						"avg_row": 34,
					},
					"field_tag_map": {
						"liquification": {
							"label": "Liquification",
							"formula": [
								{"col": "T", "row": 33},
								{"col": "U", "row": 33},
							],
						},
						"distillation": {
							"label": "Distillation",
							"formula": [
								{"col": "O", "row": 33, "scale": 1000},
								{"col": "P", "row": 33},
							],
						},
					},
				},
		},
	},
}

XLSX_ZIP_SIGNATURE = b"PK\x03\x04"
XLS_OLE_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


DERIVED_FIELDS = {
	"float_hrth": {"source": "float_zcpn", "divisor": 24}, 
}


@frappe.whitelist()
def get_field_lock_info(plant=None):
	"""Used by the client script to lock/unlock fields based on which ones
	are auto-populated from Excel for the given plant, so the field list
	only ever needs to live in PLANT_CONFIG (single source of truth).

	Returns:
	    all_fields: union of every fieldname that's Excel-mapped for ANY
	        plant/source (including derived fields) - used to reset fields
	        back to editable when switching to a plant that doesn't map them.
	    mapped_fields: the subset of those that ARE mapped for `plant` -
	        these should be locked (read-only) on the form.
	"""
	all_fields = set()
	mapped_fields = set()
	for plant_name, config in PLANT_CONFIG.items():
		for source_config in config.get("sources", {}).values():
			fieldnames = source_config.get("field_tag_map", {}).keys()
			all_fields.update(fieldnames)
			if plant_name == plant:
				mapped_fields.update(fieldnames)


	for derived_field, cfg in DERIVED_FIELDS.items():
		all_fields.add(derived_field)
		if cfg["source"] in mapped_fields:
			mapped_fields.add(derived_field)

	return {"all_fields": sorted(all_fields), "mapped_fields": sorted(mapped_fields)}


class DMRBoilerAndTurbineParameters(Document):

	def validate(self):
		excel_files = self.get(EXCEL_FILES_FIELDNAME)
		if not excel_files:
			return

		sources = self.get_plant_sources()

		self.set("min_max_avg", [])

		field_totals = {}     
		field_agg = {}       
		carry_over_totals = {} 
		missing_all = []

		for row in excel_files:
			if not row.excel_file:
				continue

			source_type = self.resolve_source_type(row, sources)
			source_config = self.get_source_config(sources, source_type, row.idx)

			ws = self.get_worksheet(row.excel_file)
			parsed = self.parse_source(ws, source_config)

			for fieldname, value in parsed["fields"].items():
				agg = source_config["field_tag_map"][fieldname].get("agg", "sum")
				field_agg[fieldname] = agg
				if agg == "avg":
					field_totals[fieldname] = value
				else:
					field_totals[fieldname] = field_totals.get(fieldname, 0) + value

			for fieldname, carry_total in parsed["carry_over"].items():
				carry_over_totals[fieldname] = carry_over_totals.get(fieldname, 0) + carry_total

			missing_all.extend(parsed["missing"])
			self.extend("min_max_avg", parsed["min_max_avg_rows"])

		for fieldname, value in field_totals.items():
			self.set(fieldname, value)

		for derived_field, cfg in DERIVED_FIELDS.items():
			source_field = cfg["source"]
			if source_field in field_totals:
				divisor = cfg.get("divisor", 1) or 1
				self.set(derived_field, field_totals[source_field] / divisor)

		if missing_all:
			frappe.msgprint(
				"Tag(s) not found in uploaded sheet(s), field(s) left unchanged: "
				+ ", ".join(missing_all)
			)

		if carry_over_totals:
			self.upsert_next_day_carry_over(carry_over_totals)


	def get_plant_sources(self):
		if not self.plant:
			frappe.throw("Plant is not set on this record — cannot determine column layout for parsing.")
		config = PLANT_CONFIG.get(self.plant)
		if not config or not config.get("sources"):
			frappe.throw(
				f"No source/tag mapping configured for plant '{self.plant}'. "
				f"Add an entry for it in PLANT_CONFIG."
			)
		return config["sources"]

	@staticmethod
	def resolve_source_type(row, sources):
		if row.source_type:
			return row.source_type
		if len(sources) == 1:

			return next(iter(sources))
		frappe.throw(
			f"Row #{row.idx} in Excel Files: source_type is required for '{row.excel_file}' "
			f"because this plant has multiple patterns configured "
			f"({', '.join(sources.keys())})."
		)

	@staticmethod
	def get_source_config(sources, source_type, row_idx):
		config = sources.get(source_type)
		if not config:
			frappe.throw(
				f"Row #{row_idx}: unknown source_type '{source_type}'. "
				f"Valid options for this plant: {', '.join(sources.keys())}."
			)
		return config


	def parse_source(self, ws, source_config):
		orientation = source_config.get("orientation", "row")
		if orientation == "column":
			return self.parse_source_column_oriented(ws, source_config)
		return self.parse_source_row_oriented(ws, source_config)

	def parse_source_row_oriented(self, ws, source_config):
		"""Tag is a ROW (found via tag_name_col); hourly values run across
		columns on that row. This is the original/default layout."""
		tag_col = column_index_from_string(source_config["tag_name_col"])
		start_col = column_index_from_string(source_config["hourly_start_col"])
		end_col = column_index_from_string(source_config["hourly_end_col"])

		mma_cols = source_config.get("min_max_avg_columns", DEFAULT_MIN_MAX_AVG_COLUMNS)
		engg_col = column_index_from_string(mma_cols["engg_units_col"])
		max_col = column_index_from_string(mma_cols["max_col"])
		max_time_col = column_index_from_string(mma_cols["max_time_col"])
		min_col = column_index_from_string(mma_cols["min_col"])
		min_time_col = column_index_from_string(mma_cols["min_time_col"])
		avg_col = column_index_from_string(mma_cols["avg_col"])

		header_row = self.find_header_row(ws, tag_col)
		if not header_row:
			frappe.throw(
				f"Could not locate header row: no cell in column {source_config['tag_name_col']} "
				f"matches '{TAG_NAME_HEADER_TEXT}'."
			)

		tag_row_map = self.build_tag_row_map(ws, header_row, tag_col)

		has_next_day = source_config.get("next_day_start_col") is not None
		if has_next_day:
			next_day_start_col = column_index_from_string(source_config["next_day_start_col"])
			next_day_end_col = (
				column_index_from_string(source_config["next_day_end_col"])
				if source_config.get("next_day_end_col") else ws.max_column
			)

		result = {"fields": {}, "missing": [], "min_max_avg_rows": [], "carry_over": {}}

		for fieldname, cfg in source_config["field_tag_map"].items():
			# Fixed-cell arithmetic fields (no tag lookup involved).
			if "formula" in cfg:
				result["fields"][fieldname] = self.compute_formula(ws, cfg["formula"])
				continue

			tag = cfg["tag"]
			row_idx = tag_row_map.get(tag.upper())
			if not row_idx:
				result["missing"].append(f"{tag} ({cfg['label']})")
				continue

			agg = cfg.get("agg", "sum")
			if agg == "avg":
				value = self.average_row_range(ws, row_idx, start_col, end_col)
			else:
				value = self.sum_row_range(ws, row_idx, start_col, end_col)
			result["fields"][fieldname] = value

			if has_next_day and agg == "sum":
				result["carry_over"][fieldname] = self.sum_row_range(
					ws, row_idx, next_day_start_col, next_day_end_col
				)

			result["min_max_avg_rows"].append({
				"parameter_name": cfg["label"],
				"field_name": fieldname,
				"engg_units": self.get_cell_str(ws, row_idx, engg_col),
				"max_value": self.get_cell_float(ws, row_idx, max_col),
				"max_value_time": self.get_cell_time(ws, row_idx, max_time_col),
				"min_value": self.get_cell_float(ws, row_idx, min_col),
				"min_value_time": self.get_cell_time(ws, row_idx, min_time_col),
				"average_value": self.get_cell_float(ws, row_idx, avg_col),
			})

		return result

	def parse_source_column_oriented(self, ws, source_config):
		"""Tag is a COLUMN (found via tag_name_row); hourly values run down
		rows in that column. Mirror image of the row-oriented layout above."""
		tag_row = source_config["tag_name_row"]
		start_row = source_config["hourly_start_row"]
		end_row = source_config["hourly_end_row"]

		mma_rows = source_config.get("min_max_avg_rows", DEFAULT_MIN_MAX_AVG_ROWS)
		engg_row = mma_rows["engg_units_row"]
		max_row = mma_rows["max_row"]
		max_time_row = mma_rows["max_time_row"]
		min_row = mma_rows["min_row"]
		min_time_row = mma_rows["min_time_row"]
		avg_row = mma_rows["avg_row"]

		needs_tag_lookup = any(
			"formula" not in cfg for cfg in source_config["field_tag_map"].values()
		)

		tag_col_map = {}
		if needs_tag_lookup:
			header_col = self.find_header_col(ws, tag_row)
			if not header_col:
				frappe.throw(
					f"Could not locate header column: no cell in row {tag_row} "
					f"matches '{TAG_NAME_HEADER_TEXT}'."
				)
			tag_col_map = self.build_tag_col_map(ws, header_col, tag_row)

		has_next_day = source_config.get("next_day_start_row") is not None
		if has_next_day:
			next_day_start_row = source_config["next_day_start_row"]
			next_day_end_row = (
				source_config["next_day_end_row"]
				if source_config.get("next_day_end_row") else ws.max_row
			)

		result = {"fields": {}, "missing": [], "min_max_avg_rows": [], "carry_over": {}}

		for fieldname, cfg in source_config["field_tag_map"].items():

			if "formula" in cfg:
				result["fields"][fieldname] = self.compute_formula(ws, cfg["formula"])
				continue

			tag = cfg["tag"]
			col_idx = tag_col_map.get(tag.upper())
			if not col_idx:
				result["missing"].append(f"{tag} ({cfg['label']})")
				continue

			agg = cfg.get("agg", "sum")
			if agg == "avg":
				value = self.average_col_range(ws, col_idx, start_row, end_row)
			else:
				value = self.sum_col_range(ws, col_idx, start_row, end_row)
			result["fields"][fieldname] = value

			if has_next_day and agg == "sum":
				result["carry_over"][fieldname] = self.sum_col_range(
					ws, col_idx, next_day_start_row, next_day_end_row
				)

			result["min_max_avg_rows"].append({
				"parameter_name": cfg["label"],
				"field_name": fieldname,
				"engg_units": self.get_cell_str(ws, engg_row, col_idx),
				"max_value": self.get_cell_float(ws, max_row, col_idx),
				"max_value_time": self.get_cell_time(ws, max_time_row, col_idx),
				"min_value": self.get_cell_float(ws, min_row, col_idx),
				"min_value_time": self.get_cell_time(ws, min_time_row, col_idx),
				"average_value": self.get_cell_float(ws, avg_row, col_idx),
			})

		return result

	# -- next-day carry-over --------------------------------------------------

	def upsert_next_day_carry_over(self, carry_over_values):
		if not self.date:
			frappe.msgprint("Date not set on this record — skipping next-day carry-over update.")
			return

		next_date = add_days(self.date, 1)
		filters = {"company": self.company, "plant": self.plant, "date": next_date}

		existing_name = frappe.db.get_value(self.doctype, filters, "name")

		if existing_name:
			for fieldname, carry_total in carry_over_values.items():
				existing_value = frappe.db.get_value(self.doctype, existing_name, fieldname) or 0
				new_value = existing_value + carry_total
				frappe.db.set_value(
					self.doctype, existing_name, fieldname, new_value, update_modified=True
				)
		else:
			new_doc = frappe.new_doc(self.doctype)
			new_doc.company = self.company
			new_doc.plant = self.plant
			new_doc.date = next_date
			for fieldname, carry_total in carry_over_values.items():
				new_doc.set(fieldname, carry_total)
			new_doc.flags.ignore_mandatory = True
			new_doc.insert(ignore_permissions=True)

	# -- file reading -----------------------------------------------------

	@staticmethod
	def get_worksheet(file_url):
		fname, fcontent = get_file(file_url)
		if not fcontent:
			frappe.throw(f"Uploaded file '{fname}' is empty or could not be read.")
		if isinstance(fcontent, str):
			fcontent = fcontent.encode("utf-8")

		head = fcontent.lstrip()[:8]

		if head.startswith(XLSX_ZIP_SIGNATURE):
			wb = openpyxl.load_workbook(BytesIO(fcontent), data_only=True)
			return wb.active

		if head.startswith(XLS_OLE_SIGNATURE):
			return DMRBoilerAndTurbineParameters.read_legacy_xls(fcontent, fname)

		frappe.throw(
			f"'{fname}' is not a recognizable Excel file (checked .xlsx and legacy .xls). "
			f"First bytes: {fcontent[:40]!r}"
		)

	@staticmethod
	def read_legacy_xls(fcontent, fname):
		try:
			import xlrd
		except ImportError:
			frappe.throw(
				"This is a legacy binary .xls file, which needs the 'xlrd' package to read. "
				"Run: bench pip install xlrd"
			)
		try:
			book = xlrd.open_workbook(file_contents=fcontent)
			sheet = book.sheet_by_index(0)
		except Exception as e:
			frappe.throw(f"Could not parse '{fname}' as a legacy .xls file: {e}")
		return _XlrdSheetAdapter(sheet)

	@staticmethod
	def find_header_row(ws, tag_col, max_scan_rows=15):
		for row in range(1, min(max_scan_rows, ws.max_row) + 1):
			val = ws.cell(row=row, column=tag_col).value
			if val and str(val).strip().lower() == TAG_NAME_HEADER_TEXT.lower():
				return row
		return None

	@staticmethod
	def build_tag_row_map(ws, header_row, tag_col):
		tag_row_map = {}
		for row in range(header_row + 1, ws.max_row + 1):
			val = ws.cell(row=row, column=tag_col).value
			if val:
				tag_row_map[str(val).strip().upper()] = row
		return tag_row_map

	@staticmethod
	def find_header_col(ws, tag_row, max_scan_cols=15):
		"""Column-oriented mirror of find_header_row: scans across a fixed
		row for the cell that reads 'Tag Name', returning its column index."""
		for col in range(1, min(max_scan_cols, ws.max_column) + 1):
			val = ws.cell(row=tag_row, column=col).value
			if val and str(val).strip().lower() == TAG_NAME_HEADER_TEXT.lower():
				return col
		return None

	@staticmethod
	def build_tag_col_map(ws, header_col, tag_row):
		"""Column-oriented mirror of build_tag_row_map: tags are read across
		columns to the right of header_col, on the fixed tag_row."""
		tag_col_map = {}
		for col in range(header_col + 1, ws.max_column + 1):
			val = ws.cell(row=tag_row, column=col).value
			if val:
				tag_col_map[str(val).strip().upper()] = col
		return tag_col_map

	@staticmethod
	def sum_col_range(ws, col_idx, start_row, end_row):
		"""Column-oriented mirror of sum_row_range: sums down a column
		instead of across a row."""
		total = 0
		for row in range(start_row, end_row + 1):
			val = ws.cell(row=row, column=col_idx).value
			if isinstance(val, (int, float)):
				total += val
			elif isinstance(val, str):
				try:
					total += float(val.strip())
				except (ValueError, TypeError):
					pass
		return total

	@staticmethod
	def average_col_range(ws, col_idx, start_row, end_row):
		"""Column-oriented mirror of average_row_range."""
		total = 0
		count = 0
		for row in range(start_row, end_row + 1):
			val = ws.cell(row=row, column=col_idx).value
			if isinstance(val, (int, float)):
				total += val
				count += 1
			elif isinstance(val, str):
				try:
					total += float(val.strip())
					count += 1
				except (ValueError, TypeError):
					pass
		return total / count if count else 0

	@staticmethod
	def sum_row_range(ws, row_idx, start_col, end_col):
		total = 0
		for col in range(start_col, end_col + 1):
			val = ws.cell(row=row_idx, column=col).value
			if isinstance(val, (int, float)):
				total += val
			elif isinstance(val, str):
				try:
					total += float(val.strip())
				except (ValueError, TypeError):
					pass
		return total

	@staticmethod
	def average_row_range(ws, row_idx, start_col, end_col):
		total = 0
		count = 0
		for col in range(start_col, end_col + 1):
			val = ws.cell(row=row_idx, column=col).value
			if isinstance(val, (int, float)):
				total += val
				count += 1
			elif isinstance(val, str):
				try:
					total += float(val.strip())
					count += 1
				except (ValueError, TypeError):
					pass
		return total / count if count else 0

	@staticmethod
	def get_cell_str(ws, row_idx, col):
		val = ws.cell(row=row_idx, column=col).value
		return str(val).strip() if val not in (None, "") else None

	@staticmethod
	def get_cell_float(ws, row_idx, col):
		val = ws.cell(row=row_idx, column=col).value
		if isinstance(val, (int, float)):
			return val
		if isinstance(val, str):
			try:
				return float(val.strip())
			except (ValueError, TypeError):
				return None
		return None

	@staticmethod
	def get_cell_time(ws, row_idx, col):
		"""Return HH:MM:SS for a time-only cell, or None if unreadable.
		openpyxl gives datetime.time/datetime.datetime; xlrd gives a raw
		day-fraction float (e.g. 0.25 == 06:00:00)."""
		val = ws.cell(row=row_idx, column=col).value
		if val is None or val == "":
			return None
		if isinstance(val, datetime.datetime):
			return val.time().strftime("%H:%M:%S")
		if isinstance(val, datetime.time):
			return val.strftime("%H:%M:%S")
		if isinstance(val, (int, float)):
			total_seconds = round(val * 86400) % 86400
			hh, rem = divmod(total_seconds, 3600)
			mm, ss = divmod(rem, 60)
			return f"{hh:02d}:{mm:02d}:{ss:02d}"
		return None

	@staticmethod
	def compute_formula(ws, terms):
		"""Sum fixed-cell terms, each optionally divided by a 'scale'.

		e.g. [{"col": "T", "row": 33}, {"col": "U", "row": 33}]
		     -> T33 + U33

		     [{"col": "O", "row": 33, "scale": 1000}, {"col": "P", "row": 33}]
		     -> (O33 / 1000) + P33

		Missing/non-numeric cells are treated as 0 rather than raising, so
		one blank term doesn't blank out the whole computed field.
		"""
		total = 0
		for term in terms:
			col_idx = column_index_from_string(term["col"])
			val = DMRBoilerAndTurbineParameters.get_cell_float(ws, term["row"], col_idx)
			total += (val or 0) / term.get("scale", 1)
		return total


class _XlrdSheetAdapter:
	def __init__(self, xlrd_sheet):
		self._sheet = xlrd_sheet
		self.max_row = xlrd_sheet.nrows
		self.max_column = xlrd_sheet.ncols

	def cell(self, row, column):
		r, c = row - 1, column - 1
		if r < 0 or r >= self._sheet.nrows or c < 0 or c >= self._sheet.ncols:
			return _CellAdapter(None)
		return _CellAdapter(self._sheet.cell_value(r, c))


class _CellAdapter:
	def __init__(self, value):
		self.value = value