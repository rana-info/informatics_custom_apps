import io
import frappe
from frappe.utils import getdate, add_days, formatdate
from collections import defaultdict
from erpnext.stock.get_item_details import get_conversion_factor
from frappe.utils import flt

DAY_START_TIME = "06:00:00"


def get_report_datetime_range(from_date, to_date, day_start_time=DAY_START_TIME):
    from_dt = f"{getdate(from_date)} {day_start_time}"
    to_dt = f"{add_days(getdate(to_date), 1)} {day_start_time}"
    return from_dt, to_dt


PLANT_FIELD = "branch"
SEGMENT_FIELD = "segment"

PRODUCTION_ITEMS = [
    ("100114", "Production of Ethanol from Maize", "A.1", "BL"),
    ("100112", "Production of Ethanol from DFG", "A.2", "BL"),
    ("100113", "Production of Ethanol from FCI Rice", "A.3", "BL"),
    ("100122", "Production of ENA from Maize", "A.4", "BL"),
    ("100120", "Production of ENA from DFG", "A.5", "BL"),
    ("100130", "Production of RS from Maize", "A.6", "BL"),
    ("100128", "Production of RS from DFG", "A.7", "BL"),
]


def get_production_item_groups():
    production_groups = [
        ("Production of IMFL", "A.8", "Case", "010205-IMFL-Mfg"),
        ("Production of by Country Liquor", "A.9", "Case", "010204-Country Liquor-Mfg"),
    ]

    result = []

    for name, code, uom, item_group in production_groups:
        item_codes = frappe.db.get_all(
            "Item",
            filters={
                "item_group": item_group,
                "disabled": 0
            },
            pluck="name"
        )

        result.append(
            (name, code, uom, item_codes)
        )

    return result

PRODUCTION_ITEM_GROUPS = get_production_item_groups()

CONSUMPTION_ITEMS = [
    ("106444", "Maize", "Quintal", "maize_opening_balance", "maize_closing_balance", "B.1", "B.6", "B.7"),
    ("100474", "DFG", "Quintal", "dfg_opening_balance", "dfg_closing_balance", "B.2", "B.8", "B.9"),
    ("106448", "FCI Surplus Rice", "Quintal", "fci_opening_balance", "fci_closing_balance", "B.3", "B.10", "B.11"),
]

ENA_CONSUMPTION_ROWS = [
    ("B.4", "ENA Mfg Consumed", "BL", "010202-ENA-Mfg"),
    ("B.5", "ENA Trd Consumed", "BL", "020201-ENA-Trd"),
]

def get_ena_consumption_rows():
    result = []

    for name, code, uom, item_group in ENA_CONSUMPTION_ROWS:
        item_codes = frappe.db.get_all(
            "Item",
            filters={
                "item_group": item_group,
                "disabled": 0
            },
            pluck="name"
        )

        result.append(
            (name, code, uom, item_codes)
        )

    return result

ENA_CONSUMPTION_ROWS = get_ena_consumption_rows()


BYPRODUCT_ITEMS = [
    ("100151", "DWGS from Maize", "D.1", "Quintal", "maize"),
    ("100149", "DWGS from DFG", "D.2", "Quintal", "dfg"),
    ("100150", "DWGS from FCI", "D.3", "Quintal", "fci"),
    ("100147", "DDGS from Maize", "D.4", "Quintal", "maize"),
    ("100145", "DDGS from DFG", "D.5", "Quintal", "dfg"),
    ("100146", "DDGS from FCI", "D.6", "Quintal", "fci"),
    ("129946", "Crude Corn Oil", "D.7", "LTR", None),
]

TECHNICAL_PARAMETER_ROWS = [
    ("alcohol_percentage", "Alcohol Percentage", "F.1", "%"),
    ("fermenter_rs", "Fermenter -RS", "F.2", "%"),
    ("fermenter_rst", "Fermenter -RST", "F.3", "%"),
    ("average_starch__maize", "Average Starch - Maize", "F.4", "%"),
    ("average_starch_dfg", "Average Starch - DFG", "F.5", "%"),
    ("average_starch_fci", "Average Starch - FCI", "F.6", "%"),
]

STOCK_ITEMS = [
    ("100114", "Stock of Ethanol from Maize", "G.2", "LTR"),
    ("100112", "Stock of Ethanol from DFG", "G.3", "LTR"),
    ("100113", "Stock of Ethanol from FCI Rice", "G.4", "LTR"),
    ("100122", "Stock of ENA from Maize", "G.5", "LTR"),
    ("100120", "Stock of ENA from DFG", "G.6", "LTR"),
    ("100130", "Stock of RS from Maize", "G.7", "LTR"),
    ("100128", "Stock of RS from DFG", "G.8", "LTR"),
]

STOCK_BYPRODUCT_GROUPS = [
    ("Stock of DDGS", "G.9", "LTR", [100147, 100145, 100146]),
    ("Stock of DWGS", "G.10", "LTR", [100151, 100149, 100150]),
    ("Stock of Crude Oil", "G.11", "LTR", [129946]),
]

BOILER_TURBINE_ITEMS = [
    ("float_zcpn", "Steam Produced", "H.1", "MT"),
    ("float_jgdk", "Steam Purchased", "H.2", "MT"),
    ("float_smgv", "Steam consumed Through PRDS", "H.3", "MT"),
    ("float_lcuw", "Steam Used Through Turbine", "H.4", "MT"),
    (None, "Sales of Steam", "H.5", "MT"),
    (None, "Short and Excess Steam", "H.6", "MT"),
    (None, "Captive Steam", "H.7", "MT"),
]

FUEL_ITEMS = [
    ("106441", "Paddy", "I.1", "MT"),
    ("106436", "Rice Husk", "I.2", "MT"),
    ("100093", "Bagasse", "I.3", "MT"),
    ("101077", "Cane Trash", "I.4", "MT"),
    ("106440", "Mustard Husk", "I.5", "MT"),
    ("106442", "Mandi Husk", "I.6", "MT"),
    ("106983", "Khudi", "I.7", "MT"),
    ("106443", "Wooden Chips", "I.8", "MT"),
]

FUEL_ITEM_GROUPS = [
    ("Coal", "I.9", "MT", [118488]),
]

BOILER_STEAM_CONSUMED_ITEMS = [
    ("liquification", "Liquification", "K.1", "MT"),
    ("distillation", "Distillation", "K.2", "MT"),
    ("msdh", "MSDH", "K.3", "MT"),
    ("evaporation", "Evaporation", "K.4", "MT"),
    ("dryer", "Dryer", "K.5", "MT"),
    ("deaerator", "Deaerator", "K.6", "MT"),
    ("other", "Other", "K.7", "MT"),
]

POWER_PERFORMANCE_ITEMS = [
    ("float_pvrh", "Power Produced", "M.1", "MW"),
    ("float_iunw", "Power Purchased", "M.2", "MW"),
    ("float_vjzq", "Power Export", "M.3", "MW"),
    ("float_dctr", "Power Sales", "M.4", "MW"),
    ("float_evng", "Captive Consumption", "M.5", "MW"),
    (None, "Short and Excess Power", "M.6", "MW"),
]

BOILER_TECHNICAL_PARAMETER_ROWS = [
    ("float_reke", "ESP Outlet Temp", "N.1", "Degree"),
    ("percent_bman", "Unburnt Ash %", "N.2", "%"),
    ("float_hrth", "Bolier Load Per Hour", "N.3", "MT/Hr"),
]

SECTION_COLORS_HEX = {
    "A": "E3F2FD",
    "B": "FFF3E0",
    "C": "E8F5E9",
    "D": "F3E5F5",
    "E": "E0F7FA",
    "F": "FCE4EC",
    "G": "FFF8E1",
    "H": "EFEBE9",
    "I": "E8EAF6",
    "J": "F1F8E9",
    "K": "E0F2F1",
    "L": "FFEBEE",
    "M": "EDE7F6",
    "N": "E1F5FE",
}


def get_fiscal_year_start(date):
    fy = frappe.db.sql("""
        select year_start_date from `tabFiscal Year`
        where %(date)s between year_start_date and year_end_date limit 1
    """, {"date": date})
    return fy[0][0] if fy else date


def build_plant_segment_conditions(plants, segments):
    conditions, values = "", {}
    if plants:
        conditions += f" and se.{PLANT_FIELD} in %(plants)s"
        values["plants"] = plants
    if segments:
        conditions += f" and sed.{SEGMENT_FIELD} in %(segments)s"
        values["segments"] = segments
    return conditions, values


def get_valid_target_uom(item_code, static_uom):
    """Returns the UOM to actually use for this item.

    If static_uom is missing, or isn't the item's stock_uom and isn't
    present in the item's UOM Conversion Detail table, falls back to the
    item's stock_uom instead of the hardcoded static UOM.
    """
    if not item_code:
        return static_uom

    stock_uom = frappe.get_cached_value("Item", item_code, "stock_uom")
    if not stock_uom:
        # Item doesn't exist / has no stock_uom - nothing sensible to fall back to.
        return static_uom

    if not static_uom or static_uom == stock_uom:
        return stock_uom

    has_conversion = frappe.db.exists(
        "UOM Conversion Detail", {"parent": item_code, "uom": static_uom}
    )
    return static_uom if has_conversion else stock_uom


def get_group_display_uom(item_codes, static_uom):
    """For a group of items sharing one report row: if every item in the
    group supports static_uom, use it. Otherwise fall back to the first
    item's stock_uom as the group's display label."""
    if not item_codes:
        return static_uom

    for code in item_codes:
        if get_valid_target_uom(code, static_uom) != static_uom:
            fallback = frappe.get_cached_value("Item", item_codes[0], "stock_uom")
            return fallback or static_uom
    return static_uom


def get_stock_to_target_factor(item_code, target_uom):
    stock_uom = frappe.get_cached_value("Item", item_code, "stock_uom")
    if not stock_uom:
        return 1
    effective_uom = get_valid_target_uom(item_code, target_uom)
    if effective_uom == stock_uom:
        return 1
    factor = (get_conversion_factor(item_code, effective_uom) or {}).get("conversion_factor") or 1
    return 1 / factor if factor else 1


def convert_qty_dict(qty_by_item, uom_map):
    return {
        code: qty * get_stock_to_target_factor(code, uom_map[code])
        for code, qty in qty_by_item.items() if code in uom_map
    }


def get_production_qty(companies, from_date, to_date, item_codes, plants=None, segments=None):
    if not item_codes:
        return {}
    from_dt, to_dt = get_report_datetime_range(from_date, to_date)
    extra_sql, extra_vals = build_plant_segment_conditions(plants, segments)
    values = {"companies": companies, "from_dt": from_dt, "to_dt": to_dt, "items": item_codes, **extra_vals}
    rows = frappe.db.sql(f"""
        select sed.item_code, sum(sed.qty * sed.conversion_factor) as qty
        from `tabStock Entry` se
        inner join `tabStock Entry Detail` sed on se.name = sed.parent
        where se.docstatus = 1 and se.stock_entry_type = 'Material Receipt'
            and se.company in %(companies)s
            and timestamp(se.posting_date, se.posting_time) >= %(from_dt)s
            and timestamp(se.posting_date, se.posting_time) < %(to_dt)s
            and sed.item_code in %(items)s {extra_sql}
        group by sed.item_code
    """, values, as_dict=1)
    return {r.item_code: r.qty or 0 for r in rows}


def get_issued_qty(companies, from_date, to_date, item_codes, plants=None, segments=None):
    if not item_codes:
        return {}
    from_dt, to_dt = get_report_datetime_range(from_date, to_date)
    extra_sql, extra_vals = build_plant_segment_conditions(plants, segments)
    values = {"companies": companies, "from_dt": from_dt, "to_dt": to_dt, "items": item_codes, **extra_vals}
    rows = frappe.db.sql(f"""
        select sed.item_code, sum(sed.qty * sed.conversion_factor) as qty
        from `tabStock Entry` se
        inner join `tabStock Entry Detail` sed on se.name = sed.parent
        where se.docstatus = 1 and se.stock_entry_type = 'Material Issue'
            and se.company in %(companies)s
            and timestamp(se.posting_date, se.posting_time) >= %(from_dt)s
            and timestamp(se.posting_date, se.posting_time) < %(to_dt)s
            and sed.item_code in %(items)s {extra_sql}
        group by sed.item_code
    """, values, as_dict=1)
    return {r.item_code: r.qty or 0 for r in rows}


def get_production_qty_by_items(companies, from_date, to_date, item_codes, target_uom, plants=None, segments=None):
    if not item_codes:
        return None
    item_codes = [str(c) for c in item_codes]
    raw = get_production_qty(companies, from_date, to_date, item_codes, plants, segments)
    converted = convert_qty_dict(raw, {c: target_uom for c in item_codes})
    return sum(converted.get(c, 0) for c in item_codes)


def get_issued_qty_by_items(companies, from_date, to_date, item_codes, target_uom, plants=None, segments=None):
    if not item_codes:
        return None
    item_codes = [str(c) for c in item_codes]
    raw = get_issued_qty(companies, from_date, to_date, item_codes, plants, segments)
    converted = convert_qty_dict(raw, {c: target_uom for c in item_codes})
    return sum(converted.get(c, 0) for c in item_codes)


def get_lab_parameter_sum(companies, date, field, plants=None):
    filters = {"company": ["in", companies], "date": getdate(date)}
    if plants:
        filters["plant"] = ["in", plants]
    values = frappe.get_all("DMR Technical Lab Parameters", filters=filters, pluck=field)
    return sum(v or 0 for v in values)


def _plucked_range_values(doctype, companies, from_date, to_date, field, plants=None):
    filters = {"company": ["in", companies], "date": ["between", [getdate(from_date), getdate(to_date)]]}
    if plants:
        filters["plant"] = ["in", plants]
    return [v for v in frappe.get_all(doctype, filters=filters, pluck=field) if v not in (None, "")]


def get_lab_parameter_avg(companies, from_date, to_date, field, plants=None):
    values = _plucked_range_values("DMR Technical Lab Parameters", companies, from_date, to_date, field, plants)
    return (sum(values) / len(values)) if values else None


def get_stock_balance_qty(companies, as_of_date, item_codes, plants=None):
    if not item_codes:
        return {}

    _, to_dt = get_report_datetime_range(as_of_date, as_of_date)

    values = {
        "companies": companies,
        "to_dt": to_dt,
        "items": item_codes,
    }

    plant_sql = ""
    if plants and frappe.db.has_column("Warehouse", "custom_branch"):
        plant_sql = " AND w.custom_branch IN %(plants)s"
        values["plants"] = plants

    rows = frappe.db.sql(f"""
        SELECT
            sle.item_code,
            sle.warehouse,
            sle.actual_qty,
            sle.qty_after_transaction,
            sle.voucher_type,
            sle.voucher_detail_no,
            sle.batch_no,
            sle.serial_no,
            sle.serial_and_batch_bundle
        FROM `tabStock Ledger Entry` sle
        INNER JOIN `tabWarehouse` w
            ON w.name = sle.warehouse
        WHERE
            sle.docstatus < 2
            AND sle.is_cancelled = 0
            AND sle.company IN %(companies)s
            AND TIMESTAMP(sle.posting_date, sle.posting_time) < %(to_dt)s
            AND sle.item_code IN %(items)s
            {plant_sql}
        ORDER BY
            sle.item_code,
            sle.warehouse,
            sle.posting_date,
            sle.posting_time,
            sle.creation
    """, values, as_dict=True)

    balance = defaultdict(float)

    for d in rows:
        qty_diff = flt(d.actual_qty)

        if (
            d.voucher_type == "Stock Reconciliation"
            and (not d.batch_no or d.serial_no or d.serial_and_batch_bundle)
        ):
            previous = balance[(d.item_code, d.warehouse)]
            qty_diff = flt(d.qty_after_transaction) - previous

        balance[(d.item_code, d.warehouse)] += qty_diff

    item_balance = defaultdict(float)

    for (item, warehouse), qty in balance.items():
        item_balance[item] += qty

    return dict(item_balance)


def get_boiler_turbine_value(companies, from_date, to_date, field, plants=None):
    if not field:
        return None
    values = _plucked_range_values("DMR Boiler And Turbine Parameters", companies, from_date, to_date, field, plants)
    return sum(values) if values else None


def get_boiler_turbine_avg(companies, from_date, to_date, field, plants=None):
    values = _plucked_range_values("DMR Boiler And Turbine Parameters", companies, from_date, to_date, field, plants)
    return (sum(values) / len(values)) if values else None


def get_steam_ratio_standards(companies, item_codes, plants=None):
    if not item_codes:
        return {}
    filters = {"parent": ["in", companies], "item": ["in", item_codes]}
    if plants:
        filters["plant"] = ["in", plants]
    rows = frappe.get_all("Steam Ratio Item", filters=filters, fields=["item", "ratio"])

    grouped = defaultdict(list)
    for r in rows:
        if r.ratio not in (None, ""):
            grouped[r.item].append(r.ratio)
    return {item: sum(vals) / len(vals) for item, vals in grouped.items() if vals}


def build_consumption_rows(companies, from_date, to_date, plants=None, segments=None):
    rows = [{"sr": "B", "label": "Consumption of Raw Material", "header": True}]

    item_codes = [c[0] for c in CONSUMPTION_ITEMS]
    issued = get_issued_qty(companies, from_date, to_date, item_codes, plants, segments)
    is_range = getdate(from_date) != getdate(to_date)

    quintal_consumed_total = 0

    consumed_rows = []
    opening_rows = []
    closing_rows = []

    for item_code, label, uom, opening_field, closing_field, consumed_sr, opening_sr, closing_sr in CONSUMPTION_ITEMS:
        display_uom = get_valid_target_uom(item_code, uom)
        factor = get_stock_to_target_factor(item_code, uom)
        opening = get_lab_parameter_sum(companies, from_date, opening_field, plants) * factor
        closing = get_lab_parameter_sum(companies, to_date, closing_field, plants) * factor
        issued_qty = (issued.get(item_code, 0) or 0) * factor
        net_consumed = opening + issued_qty - closing

        quintal_consumed_total += net_consumed

        opening_display = None if is_range else opening
        closing_display = None if is_range else closing

        consumed_rows.append({"sr": consumed_sr, "label": f"{label} Consumed ( Net of WIP)", "uom": display_uom,
                               "value": net_consumed, "item_code": item_code, "exclude_from_total": True})
        opening_rows.append({"sr": opening_sr, "label": f"Opening WIP {label}", "uom": display_uom,
                              "value": opening_display, "item_code": item_code, "exclude_from_total": True})
        closing_rows.append({"sr": closing_sr, "label": f"Closing WIP {label}", "uom": display_uom,
                              "value": closing_display, "item_code": item_code, "exclude_from_total": True})

    # sr on subheaders is a stable, non-colliding key (used as react key / excel anchor only —
    # it never appears in columns[i]["values"], so it can't clash with a real data row's sr)
    rows.append({"sr": "B.I", "label": "----------Consumed ( Net of WIP)----------", "subheader": True, "exclude_from_total": True})
    rows.extend(consumed_rows)

    rows.append({"sr": "B.II", "label": "----------Opening----------", "subheader": True, "exclude_from_total": True})
    rows.extend(opening_rows)

    rows.append({"sr": "B.III", "label": "----------Closing----------", "subheader": True, "exclude_from_total": True})
    rows.extend(closing_rows)

    rows.append({"sr": "B.Total.Quintal", "label": "Total", "uom": "Quintal",
                  "value": quintal_consumed_total, "total": True})

    for sr, label, uom, ena_item_codes in ENA_CONSUMPTION_ROWS:
        qty = get_production_qty_by_items(companies, from_date, to_date, ena_item_codes, uom, plants, segments)
        rows.append({"sr": sr, "label": label, "uom": uom, "value": qty, "item_codes": ena_item_codes})

    return rows


def _add_section_totals(rows):
    def qualifies(r):
        uom = r.get("uom")
        return not r.get("exclude_from_total") and uom and uom not in ("%", "Ratio") and "/" not in uom

    result = []
    i, n = 0, len(rows)
    while i < n:
        header_row = rows[i]
        result.append(header_row)
        if not header_row.get("header"):
            i += 1
            continue

        j = i + 1
        while j < n and not rows[j].get("header"):
            j += 1
        section_rows = rows[i + 1:j]

        last_idx_for_uom = {}
        for idx, r in enumerate(section_rows):
            if qualifies(r):
                last_idx_for_uom[r["uom"]] = idx

        seen_totals_for_uom = defaultdict(int)
        for idx, row in enumerate(section_rows):
            result.append(row)
            uom = row.get("uom")
            if not (qualifies(row) and last_idx_for_uom.get(uom) == idx):
                continue
            group = [r for r in section_rows if qualifies(r) and r.get("uom") == uom]
            if len(group) < 2:
                continue
            vals = [r.get("value") for r in group]
            note = next((r["total_note"] for r in group if r.get("total_note")), None)
            display_uom = note.format(count=len(group)) if note else uom
            seen_totals_for_uom[uom] += 1
            sr_suffix = uom if seen_totals_for_uom[uom] == 1 else f"{uom}{seen_totals_for_uom[uom]}"
            if any(isinstance(v, dict) for v in vals):
                value = {"ideal": sum((v or {}).get("ideal") or 0 for v in vals),
                          "actual": sum((v or {}).get("actual") or 0 for v in vals)}
            else:
                value = sum(v or 0 for v in vals)
            result.append({"sr": f"{header_row['sr']}.Total.{sr_suffix}", "label": "Total",
                            "uom": display_uom, "value": value, "total": True})
        i = j
    return result


def build_section_rows(companies, from_date, to_date, plants=None, segments=None):
    prod_codes = [r[0] for r in PRODUCTION_ITEMS]
    prod_uom = {code: uom for code, _, _, uom in PRODUCTION_ITEMS}
    prod = convert_qty_dict(get_production_qty(companies, from_date, to_date, prod_codes, plants, segments), prod_uom)

    rows = []

    rows.append({"sr": "A", "label": "Production of Finished Goods", "header": True})
    for code, label, sr, uom in PRODUCTION_ITEMS:
        display_uom = get_valid_target_uom(code, uom)
        rows.append({"sr": sr, "label": label, "uom": display_uom, "value": prod.get(code, 0), "item_code": code})
    for label, sr, uom, item_codes in PRODUCTION_ITEM_GROUPS:
        display_uom = get_group_display_uom(item_codes, uom) if item_codes else uom
        qty = get_production_qty_by_items(companies, from_date, to_date, item_codes, uom, plants, segments)
        rows.append({"sr": sr, "label": label, "uom": display_uom, "value": qty, "item_codes": item_codes})

    consumption_rows = build_consumption_rows(
        companies, from_date, to_date, plants, segments
    )

    rows.extend(consumption_rows)
    rows.extend(build_recovery_rows(prod, consumption_rows))

    byproduct_rows, byproduct_qty = build_byproduct_rows(companies, from_date, to_date, plants, segments)
    rows.extend(byproduct_rows)
    rows.extend(build_byproduct_recovery_rows(byproduct_qty, consumption_rows))

    rows.extend(build_technical_parameter_rows(companies, from_date, to_date, plants))
    rows.extend(build_boiler_turbine_rows(companies, from_date, to_date, plants))

    fuel_rows, fuel_qty = build_fuel_rows(companies, from_date, to_date, plants, segments)
    rows.extend(fuel_rows)
    rows.extend(build_steam_ratio_rows(companies, from_date, to_date, fuel_qty, plants))

    total_production_bl = sum(
        prod.get(code, 0) or 0
        for code, _, _, uom in PRODUCTION_ITEMS
        if uom == "BL"
    )

    rows.extend(build_steam_consumed_rows(
        companies, from_date, to_date, plants, total_production_bl
    ))
    rows.extend(build_power_performance_rows(companies, from_date, to_date, plants))
    rows.extend(build_boiler_technical_parameter_rows(companies, from_date, to_date, plants))

    rows.extend(build_stock_rows(companies, from_date, to_date, plants))

    return _add_section_totals(rows)


def _parse_list_arg(val):
    if isinstance(val, str):
        if val in ("", "null", "None", "undefined"):
            return None
        return frappe.parse_json(val)
    return val


def _build_report_payload(companies, from_date, to_date, plants=None, segments=None):
    from_date, to_date = getdate(from_date), getdate(to_date)

    date_list = []
    d = from_date
    while d <= to_date:
        date_list.append(d)
        d = add_days(d, 1)

    fy_start = get_fiscal_year_start(to_date)

    meta_rows = build_section_rows(companies, from_date, from_date, plants, segments)
    meta = [{
        "sr": r["sr"], "label": r["label"], "uom": r.get("uom"), "header": r.get("header", False),
        "item_code": r.get("item_code"), "item_codes": r.get("item_codes"),
        "total": r.get("total", False), "standard": r.get("standard"),
    } for r in meta_rows]

    columns = []
    for dt in date_list:
        rows = build_section_rows(companies, dt, dt, plants, segments)
        columns.append({"label": formatdate(dt, "dd/mm/yy"), "values": {r["sr"]: r.get("value") for r in rows}})

    todate_rows = build_section_rows(companies, fy_start, to_date, plants, segments)
    todate_label = f"To Date ({formatdate(fy_start, 'dd/mm/yy')} - {formatdate(to_date, 'dd/mm/yy')})"
    columns.append({"label": todate_label, "values": {r["sr"]: r.get("value") for r in todate_rows}})

    return {"meta": meta, "columns": columns}


def build_recovery_rows(prod, consumption_rows):
    rows = [{
        "sr": "C",
        "label": "Recovery of Finished Goods",
        "header": True,
    }]

    values = {r["sr"]: r.get("value") or 0 for r in consumption_rows if not r.get("header")}

    maize = values.get("B.1", 0)
    dfg = values.get("B.2", 0)
    fci = values.get("B.3", 0)

    ena_maize = prod.get("100122", 0)
    ena_dfg = prod.get("100120", 0)

    ethanol_maize = prod.get("100114", 0)
    ethanol_dfg = prod.get("100112", 0)
    ethanol_fci = prod.get("100113", 0)

    ena_consumed = (values.get("B.4", 0) or 0) + (values.get("B.5", 0) or 0)

    def div(a, b):
        return round(a / b, 2) if b else 0

    def pct(a, b):
        return round((a / b) * 100, 2) if b else 0

    rows.extend([
        {
            "sr": "C.1",
            "label": "Recovery from Maize",
            "uom": "BL/Quintal",
            "value": div(ena_maize + ethanol_maize, maize),
        },
        {
            "sr": "C.2",
            "label": "Recovery from DFG",
            "uom": "BL/Quintal",
            "value": div(ena_dfg + ethanol_dfg, dfg),
        },
        {
            "sr": "C.3",
            "label": "Recovery from FCI Rice",
            "uom": "BL/Quintal",
            "value": div(ethanol_fci, fci),
        },
        {
            "sr": "C.4",
            "label": "Recovery IMFL",
            "uom": "%",
            "value": pct(0, ena_consumed),
        },
        {
            "sr": "C.5",
            "label": "Recovery Country Liquor",
            "uom": "%",
            "value": pct(0, ena_consumed),
        },
    ])

    return rows


def build_byproduct_rows(companies, from_date, to_date, plants=None, segments=None):
    rows = [{"sr": "D", "label": "Production of by Products", "header": True}]

    byproduct_codes = [c[0] for c in BYPRODUCT_ITEMS]
    byproduct_uom = {code: uom for code, _, _, uom, _ in BYPRODUCT_ITEMS}
    byproduct_qty = convert_qty_dict(
        get_production_qty(companies, from_date, to_date, byproduct_codes, plants, segments),
        byproduct_uom,
    )

    for code, label, sr, uom, _raw_key in BYPRODUCT_ITEMS:
        display_uom = get_valid_target_uom(code, uom)
        rows.append({"sr": sr, "label": label, "uom": display_uom, "value": byproduct_qty.get(code, 0), "item_code": code})

    return rows, byproduct_qty


def build_byproduct_recovery_rows(byproduct_qty, consumption_rows):
    rows = [{"sr": "E", "label": "Recovery of By Products", "header": True}]

    values = {r["sr"]: r.get("value") or 0 for r in consumption_rows if not r.get("header")}

    maize = values.get("B.1", 0)
    dfg = values.get("B.2", 0)
    fci = values.get("B.3", 0)
    total_raw = maize + dfg + fci

    dwgs_maize = byproduct_qty.get("100151", 0)
    dwgs_dfg = byproduct_qty.get("100149", 0)
    dwgs_fci = byproduct_qty.get("100150", 0)
    ddgs_maize = byproduct_qty.get("100147", 0)
    ddgs_dfg = byproduct_qty.get("100145", 0)
    ddgs_fci = byproduct_qty.get("100146", 0)
    crude_oil = byproduct_qty.get("129946", 0)

    def div(a, b):
        return round(a / b, 2) if b else 0

    rows.extend([
        {"sr": "E.1", "label": "DWGS from Maize", "uom": "Quintal/Quintal", "value": div(dwgs_maize, maize)},
        {"sr": "E.2", "label": "DWGS from DFG", "uom": "Quintal/Quintal", "value": div(dwgs_dfg, dfg)},
        {"sr": "E.3", "label": "DWGS from FCI", "uom": "Quintal/Quintal", "value": div(dwgs_fci, fci)},
        {"sr": "E.4", "label": "Average DWGS", "uom": "Quintal/Quintal",
         "value": div(dwgs_maize + dwgs_dfg + dwgs_fci, total_raw)},
        {"sr": "E.5", "label": "DDGS from Maize", "uom": "Quintal/Quintal", "value": div(ddgs_maize, maize)},
        {"sr": "E.6", "label": "DDGS from DFG", "uom": "Quintal/Quintal", "value": div(ddgs_dfg, dfg)},
        {"sr": "E.7", "label": "DDGS from FCI", "uom": "Quintal/Quintal", "value": div(ddgs_fci, fci)},
        {"sr": "E.8", "label": "Average DDGS", "uom": "Quintal/Quintal",
         "value": div(ddgs_maize + ddgs_dfg + ddgs_fci, total_raw)},
        {"sr": "E.9", "label": "Crude Corn Oil", "uom": "LTR/Quintal", "value": div(crude_oil, total_raw)},
    ])

    return rows


def build_technical_parameter_rows(companies, from_date, to_date, plants=None):
    rows = [{"sr": "F", "label": "Technical Parameters", "header": True}]
    for field, label, sr, uom in TECHNICAL_PARAMETER_ROWS:
        value = get_lab_parameter_avg(companies, from_date, to_date, field, plants)
        rows.append({"sr": sr, "label": label, "uom": uom, "value": value})
    return rows


def build_stock_rows(companies, from_date, to_date, plants=None):
    rows = [{"sr": "G", "label": "Stock", "header": True}]

    wash_available = get_lab_parameter_sum(companies, to_date, "wash_available", plants)
    rows.append({"sr": "G.1", "label": "Wash Available", "uom": "LTR", "value": wash_available})

    stock_codes = [c[0] for c in STOCK_ITEMS]
    stock_uom = {code: uom for code, _, _, uom in STOCK_ITEMS}
    stock_qty = convert_qty_dict(
        get_stock_balance_qty(companies, to_date, stock_codes, plants), stock_uom
    )
    for code, label, sr, uom in STOCK_ITEMS:
        display_uom = get_valid_target_uom(code, uom)
        rows.append({"sr": sr, "label": label, "uom": display_uom, "value": stock_qty.get(code, 0), "item_code": code})

    for label, sr, uom, item_codes in STOCK_BYPRODUCT_GROUPS:
        item_codes_str = [str(c) for c in item_codes]
        display_uom = get_group_display_uom(item_codes_str, uom)
        raw_qty = get_stock_balance_qty(companies, to_date, item_codes_str, plants)
        converted = convert_qty_dict(raw_qty, {c: uom for c in item_codes_str})
        qty = sum(converted.get(c, 0) for c in item_codes_str)
        rows.append({"sr": sr, "label": label, "uom": display_uom, "value": qty, "item_codes": item_codes})

    return rows


def build_boiler_turbine_rows(companies, from_date, to_date, plants=None):
    rows = [{"sr": "H", "label": "Boiler & Turbine", "header": True}]

    fetched = {
        sr: get_boiler_turbine_value(companies, from_date, to_date, field, plants)
        for field, _, sr, _ in BOILER_TURBINE_ITEMS
    }

    supply_srs = ("H.1", "H.2")
    usage_srs = ("H.3", "H.4", "H.5", "H.6", "H.7")
    supply_total = sum(fetched.get(s, 0) or 0 for s in supply_srs)
    usage_total = sum(fetched.get(s, 0) or 0 for s in usage_srs)

    for field, label, sr, uom in BOILER_TURBINE_ITEMS:
        rows.append({"sr": sr, "label": label, "uom": uom, "value": fetched[sr], "exclude_from_total": True})
        if sr == "H.2":
            rows.append({"sr": "H.2.Total", "label": "Sub Total", "uom": "MT", "value": supply_total,
                         "total": True, "exclude_from_total": True})
        if sr == "H.7":
            rows.append({"sr": "H.7.Total", "label": "Sub Total", "uom": "MT", "value": usage_total,
                         "total": True, "exclude_from_total": True})

    rows.append({"sr": "H.Total", "label": "Total", "uom": "MT", "value": supply_total - usage_total,
                 "total": True, "exclude_from_total": True})

    return rows


def build_fuel_rows(companies, from_date, to_date, plants=None, segments=None):
    rows = [{"sr": "I", "label": "Fuel used for Boiler", "header": True}]

    fuel_codes = [code for code, _, _, _ in FUEL_ITEMS]
    fuel_uom = {code: uom for code, _, _, uom in FUEL_ITEMS}
    fuel_qty = convert_qty_dict(
        get_issued_qty(companies, from_date, to_date, fuel_codes, plants, segments), fuel_uom
    )

    for code, label, sr, uom in FUEL_ITEMS:
        display_uom = get_valid_target_uom(code, uom)
        rows.append({"sr": sr, "label": label, "uom": display_uom, "value": fuel_qty.get(code, 0), "item_code": code})

    for label, sr, uom, item_codes in FUEL_ITEM_GROUPS:
        display_uom = get_group_display_uom(item_codes, uom)
        qty = get_issued_qty_by_items(companies, from_date, to_date, item_codes, uom, plants, segments)
        fuel_qty[label] = qty
        rows.append({"sr": sr, "label": label, "uom": display_uom, "value": qty, "item_codes": item_codes})

    return rows, fuel_qty


def build_steam_ratio_rows(companies, from_date, to_date, fuel_qty, plants=None):
    rows = [{"sr": "J", "label": "Steam Raising Ratio", "header": True}]

    steam_produced = get_boiler_turbine_value(companies, from_date, to_date, "float_zcpn", plants)

    def div_ratio(a, b):
        return round(a / b, 2) if a and b else None

    fuel_codes = [code for code, _, _, _ in FUEL_ITEMS]
    standards = get_steam_ratio_standards(companies, fuel_codes, plants)

    total_fuel = 0
    for code, label, sr, uom in FUEL_ITEMS:
        qty = fuel_qty.get(code, 0) or 0
        total_fuel += qty
        rows.append({
            "sr": "J" + sr[1:],
            "label": label,
            "uom": "Ratio",
            "value": div_ratio(steam_produced, qty),
            "standard": standards.get(code),
        })

    for label, _sr, _uom, _item_codes in FUEL_ITEM_GROUPS:
        total_fuel += fuel_qty.get(label, 0) or 0

    rows.append({"sr": "J.9", "label": "Weighted Average", "uom": "Ratio",
                 "value": div_ratio(steam_produced, total_fuel)})

    return rows


def build_steam_consumed_rows(companies, from_date, to_date, plants, total_production_bl):
    rows = [{"sr": "K", "label": "Section wise Steam Consumed", "header": True, "exclude_from_total": True}]

    fetched = {
        sr: get_boiler_turbine_value(companies, from_date, to_date, field, plants)
        for field, _, sr, _ in BOILER_STEAM_CONSUMED_ITEMS
    }
    k_total = sum(fetched.get(sr, 0) or 0 for _, _, sr, _ in BOILER_STEAM_CONSUMED_ITEMS)

    for field, label, sr, uom in BOILER_STEAM_CONSUMED_ITEMS:
        rows.append({"sr": sr, "label": label, "uom": uom, "value": fetched[sr], "exclude_from_total": True})

    rows.append({"sr": "K.Total", "label": "Total", "uom": "MT", "value": k_total,
                 "total": True, "exclude_from_total": True})

    rows.append({"sr": "L", "label": "Section wise Steam Consumed per BL", "header": True, "exclude_from_total": True})

    def per_bl(mt_value):
        if mt_value is None or not total_production_bl:
            return None
        return round((mt_value * 1000) / total_production_bl, 2)

    for field, label, sr, _uom in BOILER_STEAM_CONSUMED_ITEMS:
        l_sr = "L" + sr[1:]
        rows.append({"sr": l_sr, "label": label, "uom": "KG/BL",
                     "value": per_bl(fetched[sr]), "exclude_from_total": True})

    return rows


def build_power_performance_rows(companies, from_date, to_date, plants=None):
    rows = [{"sr": "M", "label": "Power Performance", "header": True}]

    fetched = {
        sr: get_boiler_turbine_value(companies, from_date, to_date, field, plants)
        for field, _, sr, _ in POWER_PERFORMANCE_ITEMS
    }

    supply_srs = ("M.1", "M.2")
    usage_srs = ("M.3", "M.4", "M.5", "M.6")
    supply_total = sum(fetched.get(s, 0) or 0 for s in supply_srs)
    usage_total = sum(fetched.get(s, 0) or 0 for s in usage_srs)

    for field, label, sr, uom in POWER_PERFORMANCE_ITEMS:
        rows.append({"sr": sr, "label": label, "uom": uom, "value": fetched[sr], "exclude_from_total": True})
        if sr == "M.2":
            rows.append({"sr": "M.2.Total", "label": "Sub Total", "uom": "MW", "value": supply_total,
                         "total": True, "exclude_from_total": True})
        if sr == "M.6":
            rows.append({"sr": "M.6.Total", "label": "Sub Total", "uom": "MW", "value": usage_total,
                         "total": True, "exclude_from_total": True})

    rows.append({"sr": "M.Total", "label": "Total", "uom": "MW", "value": supply_total - usage_total,
                 "total": True, "exclude_from_total": True})

    return rows


def build_boiler_technical_parameter_rows(companies, from_date, to_date, plants=None):
    rows = [{"sr": "N", "label": "Technical Parameters Of Boiler & Turbine", "header": True}]
    for field, label, sr, uom in BOILER_TECHNICAL_PARAMETER_ROWS:
        value = get_boiler_turbine_avg(companies, from_date, to_date, field, plants)
        rows.append({"sr": sr, "label": label, "uom": uom, "value": value})
    return rows


def _fmt_num(val, precision=2):
    if val is None or val == "":
        return None
    try:
        n = float(val)
    except (TypeError, ValueError):
        return None
    if n == 0:
        return 0
    return round(n, 3 if abs(n) < 0.1 else precision)


def _format_scope_line(companies, plants):
    company_text = ", ".join(companies) if companies else "All Companies"
    plant_text = ", ".join(plants) if plants else "All Plants"
    return f"Company: {company_text}    |    Plant: {plant_text}"


@frappe.whitelist()
def get_report_data(companies, from_date, to_date, plants=None, segments=None):
    companies = _parse_list_arg(companies)
    plants = _parse_list_arg(plants) or None
    segments = _parse_list_arg(segments) or None
    return _build_report_payload(companies, from_date, to_date, plants, segments)


@frappe.whitelist()
def export_excel(companies, from_date, to_date, plants=None, segments=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    companies = _parse_list_arg(companies)
    plants = _parse_list_arg(plants) or None
    segments = _parse_list_arg(segments) or None
    from_date, to_date = getdate(from_date), getdate(to_date)

    payload = _build_report_payload(companies, from_date, to_date, plants, segments)
    meta, columns = payload["meta"], payload["columns"]

    wb = Workbook()
    ws = wb.active
    ws.title = "DMR"

    thin = Side(style="thin", color="AAB8C3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    n_fixed_cols = 3
    total_cols = n_fixed_cols + len(columns)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value="Daily Manufacturing Report")
    title_cell.font = Font(size=14, bold=True, color="24313B")
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    subtitle = ws.cell(row=2, column=1,
                        value=f"{formatdate(from_date, 'dd/mm/yy')} - {formatdate(to_date, 'dd/mm/yy')}  |  Values per section UOM")
    subtitle.font = Font(size=10, italic=True, color="666666")
    subtitle.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)
    scope_cell = ws.cell(row=3, column=1, value=_format_scope_line(companies, plants))
    scope_cell.font = Font(size=9.5, bold=True, color="5B7284")
    scope_cell.alignment = Alignment(horizontal="center")

    header_row_idx = 4
    headers = ["Parameters", "UOM", "Standard"] + [c["label"] for c in columns]
    header_fill = PatternFill("solid", fgColor="E8EEF4")
    for col_idx, htext in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=htext)
        cell.font = Font(bold=True, color="36474F")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.cell(row=header_row_idx, column=total_cols).fill = PatternFill("solid", fgColor="D0DFE9")

    row_idx = header_row_idx + 1
    total_fill = PatternFill("solid", fgColor="EEF3F7")
    to_date_fill = PatternFill("solid", fgColor="D7E2EC")

    for row in meta:
        if row.get("header"):
            fill = PatternFill("solid", fgColor=SECTION_COLORS_HEX.get(row["sr"], "F5F5F5"))
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols)
            cell = ws.cell(row=row_idx, column=1, value=row["label"])
            cell.font = Font(bold=True, color="1A1A1A")
            for c in range(1, total_cols + 1):
                ws.cell(row=row_idx, column=c).fill = fill
                ws.cell(row=row_idx, column=c).border = border
            row_idx += 1
            continue

        is_total = row.get("total", False)
        label_text = row["label"] if is_total else f"{row['sr']}  {row['label']}"
        if row.get("item_code"):
            label_text += f" ({row['item_code']})"
        elif row.get("item_codes"):
            label_text += f" ({', '.join(str(c) for c in row['item_codes'])})"

        ws.cell(row=row_idx, column=1, value=label_text).font = Font(bold=True, color="1F2B34")
        ws.cell(row=row_idx, column=2, value=row.get("uom") or "").alignment = Alignment(horizontal="center")

        std = _fmt_num(row.get("standard"))
        std_cell = ws.cell(row=row_idx, column=3, value=std if std is not None else None)
        std_cell.alignment = Alignment(horizontal="right")
        if std is not None:
            std_cell.number_format = "#,##0.00"

        for i, col in enumerate(columns):
            val = col["values"].get(row["sr"])
            cell = ws.cell(row=row_idx, column=n_fixed_cols + 1 + i)
            cell.border = border
            num = _fmt_num(val)
            cell.value = num if num is not None else "-"
            if num is not None:
                cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right")
            if i == len(columns) - 1:
                cell.fill = to_date_fill if not is_total else PatternFill("solid", fgColor="D7E2EC")

        if is_total:
            for c in range(1, total_cols + 1):
                existing = ws.cell(row=row_idx, column=c)
                existing.font = Font(bold=True, color=existing.font.color)
                if c != total_cols:
                    existing.fill = total_fill

        for c in range(1, total_cols + 1):
            ws.cell(row=row_idx, column=c).border = border

        row_idx += 1

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    for i in range(len(columns)):
        col_letter = get_column_letter(n_fixed_cols + 1 + i)
        ws.column_dimensions[col_letter].width = 22 if i == len(columns) - 1 else 16

    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=n_fixed_cols + 1).coordinate

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    frappe.response["filename"] = f"DMR_{from_date}_to_{to_date}.xlsx"
    frappe.response["filecontent"] = buf.getvalue()
    frappe.response["type"] = "binary"


def _render_report_html(payload, companies, plants, from_date, to_date):
    meta, columns = payload["meta"], payload["columns"]
    escape = frappe.utils.escape_html

    head_cells = "<th>Parameters</th><th>UOM</th><th>Standard</th>"
    for i, col in enumerate(columns):
        style = ' style="background:#D0DFE9;"' if i == len(columns) - 1 else ""
        head_cells += f"<th{style}>{escape(col['label'])}</th>"

    body_rows = []
    for row in meta:
        if row.get("header"):
            color = SECTION_COLORS_HEX.get(row["sr"], "F5F5F5")
            colspan = 2 + len(columns)
            body_rows.append(
                f'<tr style="background:#{color};"><td style="font-weight:bold;">{escape(row["label"])}</td>'
                f'<td colspan="{colspan}"></td></tr>')
            continue

        is_total_row = row.get("total")
        row_style = "background:#EEF3F7;font-weight:bold;" if is_total_row else ""
        label = escape(row["label"]) if is_total_row else f"{row['sr']}&nbsp;&nbsp;{escape(row['label'])}"
        if row.get("item_code"):
            label += f' <span style="color:#8ea0ac;font-size:8px;">({escape(row["item_code"])})</span>'
        elif row.get("item_codes"):
            codes = ", ".join(escape(str(c)) for c in row["item_codes"])
            label += f' <span style="color:#8ea0ac;font-size:8px;">({codes})</span>'

        cells = f'<tr style="{row_style}"><td style="text-align:left;font-weight:bold;">{label}</td>'
        cells += f'<td style="text-align:center;">{escape(row.get("uom") or "")}</td>'
        std = _fmt_num(row.get("standard"))
        cells += f'<td style="text-align:right;">{"" if std is None else f"{std:,.2f}"}</td>'

        for i, col in enumerate(columns):
            val = col["values"].get(row["sr"])
            is_last = i == len(columns) - 1
            td_style = "text-align:right;" + ("background:#D7E2EC;" if is_last else "")
            num = _fmt_num(val)
            text = "-" if num is None else f"{num:,.2f}"
            cells += f'<td style="{td_style}">{text}</td>'
        cells += "</tr>"
        body_rows.append(cells)

    date_range = f"{formatdate(from_date, 'dd/mm/yy')} - {formatdate(to_date, 'dd/mm/yy')}"
    scope_line = escape(_format_scope_line(companies, plants))
    return f"""
    <html><head><style>
        body {{ font-family: Arial, sans-serif; font-size: 8px; }}
        h2 {{ text-align:center; margin-bottom:2px; color:#24313B; }}
        p.subtitle {{ text-align:center; color:#666; margin-top:0; margin-bottom:2px; }}
        p.scope {{ text-align:center; color:#5B7284; font-weight:bold; margin-top:0; margin-bottom:10px; }}
        table {{ width:100%; border-collapse:collapse; }}
        th, td {{ border:1px solid #AAB8C3; padding:3px 5px; white-space:nowrap; }}
        th {{ background:#E8EEF4; text-align:right; color:#36474F; }}
        th:first-child, td:first-child {{ text-align:left; min-width:160px; }}
    </style></head><body>
        <h2>Daily Manufacturing Report</h2>
        <p class="subtitle">{date_range} &mdash; Values per section UOM</p>
        <p class="scope">{scope_line}</p>
        <table><thead><tr>{head_cells}</tr></thead><tbody>{''.join(body_rows)}</tbody></table>
    </body></html>
    """


@frappe.whitelist()
def export_pdf(companies, from_date, to_date, plants=None, segments=None):
    from frappe.utils.pdf import get_pdf

    companies = _parse_list_arg(companies)
    plants = _parse_list_arg(plants) or None
    segments = _parse_list_arg(segments) or None
    from_date, to_date = getdate(from_date), getdate(to_date)

    payload = _build_report_payload(companies, from_date, to_date, plants, segments)
    html = _render_report_html(payload, companies, plants, from_date, to_date)
    pdf_bytes = get_pdf(html, {"orientation": "Landscape", "page-size": "A3", "margin-top": "10mm",
                                "margin-bottom": "10mm", "margin-left": "5mm", "margin-right": "5mm"})

    frappe.response["filename"] = f"DMR_{from_date}_to_{to_date}.pdf"
    frappe.response["filecontent"] = pdf_bytes
    frappe.response["type"] = "download"


@frappe.whitelist()
def get_plant_options(companies=None):
    companies = _parse_list_arg(companies)

    if frappe.db.has_column("Branch", "company"):
        filters = {"company": ["in", companies]} if companies else {}
        return frappe.get_all("Branch", filters=filters, pluck="name", order_by="name")

    if not frappe.db.has_column("Stock Entry", PLANT_FIELD):
        return []
    condition = "and company in %(companies)s" if companies else ""
    plants = frappe.db.sql(f"""
        select distinct {PLANT_FIELD} as plant from `tabStock Entry`
        where {PLANT_FIELD} is not null and {PLANT_FIELD} != '' {condition}
        order by {PLANT_FIELD}
    """, {"companies": companies} if companies else {}, as_dict=1)
    return [p.plant for p in plants]


@frappe.whitelist()
def get_segment_options():
    filters = {"is_group": 0} if frappe.db.has_column("Segment", "is_group") else {}
    return frappe.get_all("Segment", filters=filters, pluck="name", order_by="name")