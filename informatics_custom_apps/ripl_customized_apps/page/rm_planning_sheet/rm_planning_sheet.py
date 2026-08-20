import json

import frappe
from frappe.utils import flt
from frappe.utils.xlsxutils import make_xlsx


ITEM_MAP = {
    "DFG": {"fg": "100112", "rm": "106446"},
    "Maize": {"fg": "100114", "rm": "106444"},
    "FCI": {"fg": "100113", "rm": "106448"},
}

PRODUCTS = ["DFG", "Maize", "FCI"]

# UOM codes that should be displayed under a friendlier / corrected label.
# The underlying quantities are unaffected - this only changes what label
# is shown to the user (e.g. system UOM "KLR" is shown as "KL").
UOM_DISPLAY_MAP = {
    "KLR": "KL",
}


def display_uom(uom):
    return UOM_DISPLAY_MAP.get(uom, uom)


def pos(value):
    return value if value > 0 else 0


def get_plant_warehouses(plant):
    if not plant:
        return []
    return frappe.get_all(
        "Warehouse",
        filters={"custom_branch": plant, "is_group": 0},
        pluck="name",
    )


def get_stock_uom(item_code):
    return frappe.db.get_value("Item", item_code, "stock_uom") or ""


def get_highest_uom_and_factor(item_code):
    row = frappe.db.get_value(
        "UOM Conversion Detail",
        {"parent": item_code, "parenttype": "Item"},
        ["uom", "conversion_factor"],
        order_by="conversion_factor desc",
        as_dict=True,
    )
    if row and flt(row.conversion_factor):
        return row.uom, flt(row.conversion_factor)
    return get_stock_uom(item_code), 1


def get_dispatch_qty(item_code, warehouses, from_date, to_date):
    if not warehouses:
        return 0
    return flt(frappe.db.sql("""
        select sum(dni.stock_qty)
        from `tabDelivery Note Item` dni
        inner join `tabDelivery Note` dn on dn.name = dni.parent
        where dn.docstatus = 1
            and dni.item_code = %(item_code)s
            and dni.warehouse in %(warehouses)s
            and dn.posting_date between %(from_date)s and %(to_date)s
    """, {
        "item_code": item_code,
        "warehouses": warehouses,
        "from_date": from_date,
        "to_date": to_date,
    })[0][0] or 0)


def get_stock_qty(item_code, warehouses, to_date=None):
    if not warehouses:
        return 0
    to_date = to_date or frappe.utils.today()

    rows = frappe.db.sql("""
        select warehouse, qty_after_transaction
        from (
            select
                warehouse,
                qty_after_transaction,
                row_number() over (
                    partition by warehouse
                    order by posting_date desc, posting_time desc, creation desc
                ) as rn
            from `tabStock Ledger Entry`
            where item_code = %(item_code)s
                and warehouse in %(warehouses)s
                and is_cancelled = 0
                and posting_date <= %(to_date)s
        ) ranked
        where rn = 1
    """, {"item_code": item_code, "warehouses": warehouses, "to_date": to_date}, as_dict=True)

    return flt(sum(flt(r.qty_after_transaction) for r in rows))


def get_avg_rate(item_code, warehouses, to_date=None):
    if not warehouses:
        return 0
    to_date = to_date or frappe.utils.today()

    rows = frappe.db.sql("""
        select warehouse, qty_after_transaction, stock_value
        from (
            select
                warehouse,
                qty_after_transaction,
                stock_value,
                row_number() over (
                    partition by warehouse
                    order by posting_date desc, posting_time desc, creation desc
                ) as rn
            from `tabStock Ledger Entry`
            where item_code = %(item_code)s
                and warehouse in %(warehouses)s
                and is_cancelled = 0
                and posting_date <= %(to_date)s
        ) ranked
        where rn = 1
    """, {"item_code": item_code, "warehouses": warehouses, "to_date": to_date}, as_dict=True)

    total_qty = sum(flt(r.qty_after_transaction) for r in rows if flt(r.qty_after_transaction) > 0)
    total_value = sum(flt(r.stock_value) for r in rows if flt(r.qty_after_transaction) > 0)
    return flt(total_value / total_qty) if total_qty else 0


def get_sauda_in_hand(item_code, warehouses):
    """Total contracted (Sauda / Purchase Order) quantity, regardless of
    how much of it has already been delivered."""
    if not warehouses:
        return 0
    return flt(frappe.db.sql("""
        select sum(poi.stock_qty)
        from `tabPurchase Order Item` poi
        inner join `tabPurchase Order` po on po.name = poi.parent
        where po.docstatus = 1
            and po.status not in ('Closed', 'Completed')
            and poi.item_code = %(item_code)s
            and poi.warehouse in %(warehouses)s
    """, {"item_code": item_code, "warehouses": warehouses})[0][0] or 0)


def get_sauda_not_delivered(item_code, warehouses):
    """Portion of contracted (Sauda / Purchase Order) quantity that is
    still pending delivery."""
    if not warehouses:
        return 0
    return flt(frappe.db.sql("""
        select sum(poi.stock_qty - poi.received_qty)
        from `tabPurchase Order Item` poi
        inner join `tabPurchase Order` po on po.name = poi.parent
        where po.docstatus = 1
            and po.status not in ('Closed', 'Completed')
            and poi.item_code = %(item_code)s
            and poi.warehouse in %(warehouses)s
    """, {"item_code": item_code, "warehouses": warehouses})[0][0] or 0)


def _compute_for_plant(plant, ethanol_supply_year, daily_capacity):
    filters = {"plant": plant}
    if ethanol_supply_year:
        filters["ethanol_supply_year"] = ethanol_supply_year

    alloc_name = frappe.db.get_value(
        "Ethanol Allocation", filters, "name", order_by="creation desc"
    )
    if not alloc_name:
        return None

    alloc_doc = frappe.get_doc("Ethanol Allocation", alloc_name)
    warehouses = get_plant_warehouses(plant)
    today = frappe.utils.today()

    fg_uom_factor = {p: get_highest_uom_and_factor(ITEM_MAP[p]["fg"]) for p in PRODUCTS}
    rm_uom_factor = {p: get_highest_uom_and_factor(ITEM_MAP[p]["rm"]) for p in PRODUCTS}
    fg_uom = display_uom(fg_uom_factor["DFG"][0])
    rm_uom = display_uom(rm_uom_factor["DFG"][0])

    quarters = {}
    for row in alloc_doc.allocation:
        q = quarters.setdefault(row.quarter, {"DFG": 0, "Maize": 0, "FCI": 0})
        q["DFG"] += flt(row.ethanol_from_dfg)
        q["Maize"] += flt(row.ethanol_from_maize)
        q["FCI"] += flt(row.ethanol_from_fci)

    total_allocation = {p: sum(q[p] for q in quarters.values()) for p in PRODUCTS}

    recovery = {"DFG": 0, "Maize": 0, "FCI": 0}
    if alloc_doc.recovery:
        r = alloc_doc.recovery[0]
        recovery = {"DFG": flt(r.dfg), "Maize": flt(r.maize), "FCI": flt(r.fci)}

    def conv(d, factor_map, precision=3):
        return {
            p: flt(d[p] / factor_map[p][1], precision) if factor_map[p][1] else flt(d[p], precision)
            for p in PRODUCTS
        }

    def conv_quarterdict(qd, factor_map):
        return {q: conv(vals, factor_map) for q, vals in qd.items()}

    dispatch_quarters = {}
    total_dispatch = {p: 0 for p in PRODUCTS}
    if alloc_doc.ethanol_supply_year:
        esy = frappe.get_doc("Ethanol Supply Year", alloc_doc.ethanol_supply_year)
        for row in esy.ethanol_supply_quarter:
            if not (row.start_date and row.end_date):
                continue
            dispatch_quarters[row.quarter] = {}
            for p in PRODUCTS:
                qty = get_dispatch_qty(
                    ITEM_MAP[p]["fg"], warehouses, row.start_date, row.end_date
                )
                dispatch_quarters[row.quarter][p] = qty
                total_dispatch[p] += qty
    dispatch_quarters = conv_quarterdict(dispatch_quarters, fg_uom_factor)
    total_dispatch = conv(total_dispatch, fg_uom_factor)

    stock_in_hand = {p: get_stock_qty(ITEM_MAP[p]["fg"], warehouses, today) for p in PRODUCTS}
    stock_in_hand = conv(stock_in_hand, fg_uom_factor)

    net_pending = {
        p: pos(total_allocation[p] - total_dispatch[p] - stock_in_hand[p]) for p in PRODUCTS
    }
    pending_dispatch = {
        p: pos(total_allocation[p] - total_dispatch[p]) for p in PRODUCTS
    }

    # "Qty of RM required" expressed in the ethanol (FG) highest-UOM scale
    # (previously labelled "KLR" - now shown/treated as plain "KL").
    qty_rm_required_kl = {
        p: flt(net_pending[p] / recovery[p], 0) if recovery[p] else 0 for p in PRODUCTS
    }
    qty_rm_required = {
        p: flt(qty_rm_required_kl[p] * rm_uom_factor[p][1], 3) for p in PRODUCTS
    }

    rm_at_factory = {p: get_stock_qty(ITEM_MAP[p]["rm"], warehouses, today) for p in PRODUCTS}

    sauda_in_hand = {p: get_sauda_in_hand(ITEM_MAP[p]["rm"], warehouses) for p in PRODUCTS}
    sauda_not_delivered = {p: get_sauda_not_delivered(ITEM_MAP[p]["rm"], warehouses) for p in PRODUCTS}

    net_qty_purchase = {
        p: pos(qty_rm_required[p] - rm_at_factory[p] - sauda_not_delivered[p]) for p in PRODUCTS
    }

    rate_rm = {p: get_avg_rate(ITEM_MAP[p]["rm"], warehouses, today) for p in PRODUCTS}

    value_rm = {p: flt(net_qty_purchase[p] * rate_rm[p] / 100000, 2) for p in PRODUCTS}

    def total_row(d):
        return sum(d.values())

    value_rm_total = total_row(value_rm)

    # RM quantities are rounded off to the nearest Quintal (whole number),
    # rather than carrying decimal precision.
    rm_at_factory = conv(rm_at_factory, rm_uom_factor, precision=0)
    sauda_in_hand = conv(sauda_in_hand, rm_uom_factor, precision=0)
    sauda_not_delivered = conv(sauda_not_delivered, rm_uom_factor, precision=0)
    net_qty_purchase = conv(net_qty_purchase, rm_uom_factor, precision=0)
    qty_rm_required = conv(qty_rm_required, rm_uom_factor, precision=0)

    rate_rm = {p: flt(rate_rm[p] * rm_uom_factor[p][1], 2) for p in PRODUCTS}

    # No. of days of finished-goods stock in hand, covered at full plant
    # production capacity. Daily capacity is a mandatory user-entered
    # figure (KL/day) - no historical auto-calculation is used.
    daily_capacity = flt(daily_capacity)
    stock_in_hand_total = total_row(stock_in_hand)
    days_in_hand = flt(stock_in_hand_total / daily_capacity, 1) if daily_capacity else 0

    totals = {
        "total_allocation": total_row(total_allocation),
        "total_dispatch": total_row(total_dispatch),
        "stock_in_hand": stock_in_hand_total,
        "pending_dispatch": total_row(pending_dispatch),
        "net_pending": total_row(net_pending),
        "qty_rm_required": total_row(qty_rm_required),
        "rm_at_factory": total_row(rm_at_factory),
        "sauda_in_hand": total_row(sauda_in_hand),
        "sauda_not_delivered": total_row(sauda_not_delivered),
        "net_qty_purchase": total_row(net_qty_purchase),
        "value_rm": value_rm_total,
        "days_in_hand": days_in_hand,
    }

    return {
        "plant": plant,
        "ethanol_supply_year": alloc_doc.ethanol_supply_year,
        "fg_uom": fg_uom,
        "rm_uom": rm_uom,
        "quarters": quarters,
        "total_allocation": total_allocation,
        "dispatch_quarters": dispatch_quarters,
        "total_dispatch": total_dispatch,
        "stock_in_hand": stock_in_hand,
        "pending_dispatch": pending_dispatch,
        "net_pending": net_pending,
        "recovery": recovery,
        "qty_rm_required": qty_rm_required,
        "rm_at_factory": rm_at_factory,
        "sauda_in_hand": sauda_in_hand,
        "sauda_not_delivered": sauda_not_delivered,
        "net_qty_purchase": net_qty_purchase,
        "rate_rm": rate_rm,
        "value_rm": value_rm,
        "daily_capacity": daily_capacity,
        "days_in_hand": days_in_hand,
        "totals": totals,
    }


@frappe.whitelist()
def get_active_supply_year():
    today = frappe.utils.today()
    active = frappe.db.get_value(
        "Ethanol Supply Year",
        {"disabled": 0, "year_start_date": ["<=", today], "year_end_date": [">=", today]},
        "name",
    )
    if active:
        return active
    return frappe.db.get_value(
        "Ethanol Supply Year", {"disabled": 0}, "name", order_by="year_start_date desc"
    )


@frappe.whitelist()
def get_planning_data(plants, ethanol_supply_year=None, daily_capacity=None):
    if isinstance(plants, str):
        plants = json.loads(plants)

    if daily_capacity in (None, ""):
        frappe.throw("Daily Capacity is mandatory. Please enter the plant's full production capacity.")
    daily_capacity = flt(daily_capacity)
    if daily_capacity <= 0:
        frappe.throw("Daily Capacity must be a positive number.")

    if not ethanol_supply_year:
        ethanol_supply_year = get_active_supply_year()

    results = []
    for plant in plants:
        data = _compute_for_plant(plant, ethanol_supply_year, daily_capacity)
        if data:
            results.append(data)
        else:
            results.append({"plant": plant, "no_data": True})

    return {"ethanol_supply_year": ethanol_supply_year, "plants": results}


import io
import json

import frappe
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@frappe.whitelist()
def export_planning_excel(plants, ethanol_supply_year=None, daily_capacity=None):
	if isinstance(plants, str):
		plants = json.loads(plants)

	data = get_planning_data(plants=plants, ethanol_supply_year=ethanol_supply_year, daily_capacity=daily_capacity)

	wb = Workbook()
	wb.remove(wb.active)

	header_fill = PatternFill("solid", fgColor="F4F6FA")
	plant_fill = PatternFill("solid", fgColor="EEF2F9")
	allocation_fill = PatternFill("solid", fgColor="FFFBE0")
	dispatch_fill = PatternFill("solid", fgColor="F2F9FF")
	total_fill = PatternFill("solid", fgColor="EEF1F6")
	net_fill = PatternFill("solid", fgColor="FFF0E8")
	purchase_fill = PatternFill("solid", fgColor="EAF7EC")

	thin = Side(style="thin", color="E3E6EC")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)

	columns = ["Particulars", "UOM", "DFG", "Maize", "FCI", "Total"]

	used_names = set()

	def sheet_name_for(plant):
		name = (plant or "Plant").strip()
		for ch in ["\\", "/", "*", "?", ":", "[", "]"]:
			name = name.replace(ch, "")
		name = name[:31] or "Plant"

		base = name
		i = 1
		while name in used_names:
			suffix = f" ({i})"
			name = base[: 31 - len(suffix)] + suffix
			i += 1
		used_names.add(name)
		return name

	def write_row(ws, row_idx, label, vals, uom="", fill=None, bold=False, is_total_meaningful=True):
		vals = vals or {}
		dfg = vals.get("DFG", 0) or 0
		maize = vals.get("Maize", 0) or 0
		fci = vals.get("FCI", 0) or 0
		total = (dfg + maize + fci) if is_total_meaningful else ""

		row_values = [label, uom, dfg, maize, fci, total]
		for col_idx, val in enumerate(row_values, start=1):
			cell = ws.cell(row=row_idx, column=col_idx, value=val)
			cell.border = border
			if col_idx in (1, 2):
				cell.alignment = Alignment(horizontal="left")
			else:
				cell.alignment = Alignment(horizontal="right")
				if val != "":
					cell.number_format = "#,##0.00"
			if fill:
				cell.fill = fill
			if bold:
				cell.font = Font(bold=True)

	def write_single_row(ws, row_idx, label, value, uom="", fill=None, bold=False, number_format="#,##0.0"):
		row_values = [label, uom, "", "", "", value]
		for col_idx, val in enumerate(row_values, start=1):
			cell = ws.cell(row=row_idx, column=col_idx, value=val)
			cell.border = border
			if col_idx in (1, 2):
				cell.alignment = Alignment(horizontal="left")
			else:
				cell.alignment = Alignment(horizontal="right")
				if val != "":
					cell.number_format = number_format
			if fill:
				cell.fill = fill
			if bold:
				cell.font = Font(bold=True)

	for plant_data in data.get("plants", []):
		plant = plant_data.get("plant") or "Plant"
		ws = wb.create_sheet(sheet_name_for(plant))

		ws.merge_cells("A1:F1")
		header_cell = ws["A1"]
		header_cell.value = plant
		header_cell.font = Font(bold=True, size=13)
		header_cell.fill = plant_fill
		header_cell.alignment = Alignment(horizontal="left", vertical="center")
		ws.row_dimensions[1].height = 22

		if plant_data.get("no_data"):
			ws.merge_cells("A2:F2")
			cell = ws["A2"]
			cell.value = "No Ethanol Allocation found for this plant / supply year."
			cell.font = Font(italic=True, color="8A94A6")
			ws.column_dimensions["A"].width = 55
			continue

		fg_uom = plant_data.get("fg_uom") or "KL"
		rm_uom = plant_data.get("rm_uom") or "Quintal"

		for col_idx, col_name in enumerate(columns, start=1):
			cell = ws.cell(row=2, column=col_idx, value=col_name)
			cell.font = Font(bold=True)
			cell.fill = header_fill
			cell.alignment = Alignment(horizontal="center")
			cell.border = border

		row_idx = 3

		for q, vals in (plant_data.get("quarters") or {}).items():
			write_row(ws, row_idx, f"{q} Allocation", vals, uom=fg_uom, fill=allocation_fill)
			row_idx += 1

		write_row(ws, row_idx, "Total allocation (A)", plant_data.get("total_allocation"),
			uom=fg_uom, fill=total_fill, bold=True)
		row_idx += 1

		for q, vals in (plant_data.get("dispatch_quarters") or {}).items():
			write_row(ws, row_idx, f"{q} Dispatch", vals, uom=fg_uom, fill=dispatch_fill)
			row_idx += 1

		write_row(ws, row_idx, "Total dispatch (B)", plant_data.get("total_dispatch"),
			uom=fg_uom, fill=total_fill, bold=True)
		row_idx += 1

		write_row(ws, row_idx, "Pending Dispatch (A-B)", plant_data.get("pending_dispatch"),
			uom=fg_uom, fill=net_fill, bold=True)
		row_idx += 1


		write_row(ws, row_idx, "Total stock in hand (C)", plant_data.get("stock_in_hand"),
			uom=fg_uom, fill=total_fill, bold=True)
		row_idx += 1

		write_single_row(ws, row_idx, "No. of Days in Hand (Full Capacity)",
			plant_data.get("days_in_hand", 0), uom="Days", fill=net_fill, bold=True)
		row_idx += 1

		write_row(ws, row_idx, "Net pending production (A-B-C)", plant_data.get("net_pending"),
			uom=fg_uom, fill=net_fill, bold=True)
		row_idx += 1

		write_row(ws, row_idx, "Recovery", plant_data.get("recovery"), uom="%",
			is_total_meaningful=False)
		row_idx += 1

		write_row(ws, row_idx, "Qty of RM required", plant_data.get("qty_rm_required"),
			uom=rm_uom, fill=total_fill, bold=True)
		row_idx += 1

		write_row(ws, row_idx, "RM at factory", plant_data.get("rm_at_factory"), uom=rm_uom)
		row_idx += 1

		write_row(ws, row_idx, "Sauda in hand", plant_data.get("sauda_in_hand"), uom=rm_uom)
		row_idx += 1

		write_row(ws, row_idx, "Sauda not delivered", plant_data.get("sauda_not_delivered"), uom=rm_uom)
		row_idx += 1

		write_row(ws, row_idx, "Net qty need to purchase", plant_data.get("net_qty_purchase"),
			uom=rm_uom, fill=total_fill, bold=True)
		row_idx += 1

		write_row(ws, row_idx, "Rate of RM", plant_data.get("rate_rm"), uom=f"₹/{rm_uom}",
			is_total_meaningful=False)
		row_idx += 1

		write_row(ws, row_idx, "Value of RM needs to purchase (Lakhs)", plant_data.get("value_rm"),
			uom="Lakhs", fill=purchase_fill, bold=True)
		row_idx += 1

		ws.column_dimensions["A"].width = 38
		ws.column_dimensions["B"].width = 10
		for col in ["C", "D", "E", "F"]:
			ws.column_dimensions[col].width = 14
		ws.freeze_panes = "A3"

	if not wb.sheetnames:
		wb.create_sheet("RM Planning - Ethanol Tender")

	buffer = io.BytesIO()
	wb.save(buffer)

	frappe.response["filename"] = "RM_Planning_Ethanol_Tender.xlsx"
	frappe.response["filecontent"] = buffer.getvalue()
	frappe.response["type"] = "binary"